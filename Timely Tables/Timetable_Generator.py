import random
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side

class TimetableGenerator:
    def __init__(self):
        self.subject_entries = []
        self.faculty_entries = []
        self.lab_subject_entries = []
        self.lab_faculty_entries = []
        self.num_subjects = 0
        self.num_lab_sessions = 0
        self.timetables = {}

    def get_input(self):
        # Division name input
        num_divisions = int(input("Enter Number of Divisions: "))
        division_names_input = input("Enter Division Names (comma-separated): ")
        division_names = [name.strip() for name in division_names_input.split(',')]

        # Slot input
        all_slots_input = input("Enter All Time Slots (comma-separated): ")
        lecture_slots_input = input("Enter Lecture Time Slots (comma-separated): ")
        lunch_slots_input = input("Enter Lunch Break Time Slots (comma-separated): ")

        all_slots = [slot.strip() for slot in all_slots_input.split(",")]
        lecture_slots = [slot.strip() for slot in lecture_slots_input.split(",")]
        lunch_slots = [slot.strip() for slot in lunch_slots_input.split(",")]

        # Subject and Faculty input section
        while True:
            subject_name = input("Enter Subject (or 'done' to finish): ")
            if subject_name.lower() == 'done':
                break
            faculty_names = input("Enter Faculty for this Subject (comma-separated): ")
            self.subject_entries.append(subject_name)
            self.faculty_entries.append([name.strip() for name in faculty_names.split(',')])
            self.num_subjects += 1

        # Lab Sessions input
        while True:
            lab_subject = input("Enter Lab Subject (or 'done' to finish): ")
            if lab_subject.lower() == 'done':
                break
            lab_faculty_names = input("Enter Lab Faculty (comma-separated): ")
            self.lab_subject_entries.append(lab_subject)
            self.lab_faculty_entries.append([name.strip() for name in lab_faculty_names.split(',')])
            self.num_lab_sessions += 1

        return num_divisions, division_names, all_slots, lecture_slots, lunch_slots

    def generate_timetable(self, num_divisions, division_names, all_slots, lecture_slots, lunch_slots):
        # Initialize timetables for each division
        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
        timetables = {division: pd.DataFrame(index=days, columns=all_slots) for division in division_names}

        # Track assigned lectures to avoid faculty overlaps
        faculty_schedule = {day: {slot: [] for slot in all_slots} for day in days}

        for division in division_names:
            for day in days:
                used_subjects = set()
                lab_session_scheduled = False

                # Assign one lecture per subject for each day
                for slot in lecture_slots:
                    if len(used_subjects) < len(self.subject_entries):
                        available_subjects = [subj for subj in self.subject_entries if subj not in used_subjects]
                        if available_subjects:
                            selected_subject = random.choice(available_subjects)
                            selected_faculty = random.choice(self.faculty_entries[self.subject_entries.index(selected_subject)])
                            if selected_faculty not in faculty_schedule[day][slot]:
                                timetables[division].at[day, slot] = f"{selected_subject}\n{selected_faculty}"
                                faculty_schedule[day][slot].append(selected_faculty)
                                used_subjects.add(selected_subject)

                # Schedule one two-hour lecture once a week
                if day == 'Monday':  # Example: Schedule on Monday
                    if lecture_slots:  # Ensure there are available slots
                        selected_subject = random.choice(self.subject_entries)
                        selected_faculty = random.choice(self.faculty_entries[self.subject_entries.index(selected_subject)])
                        # Assign to first two available slots in the lecture_slots
                        for i in range(len(lecture_slots) - 1):
                            if lecture_slots[i] in all_slots and lecture_slots[i + 1] in all_slots:
                                timetables[division].at[day, lecture_slots[i]] = f"{selected_subject}\n{selected_faculty}"
                                timetables[division].at[day, lecture_slots[i + 1]] = f"{selected_subject}\n{selected_faculty}"
                                break

                # Assign lab sessions in 2-hour blocks without discontinuity
                lab_slots = [(all_slots[i], all_slots[i + 1]) for i in range(len(all_slots) - 1)]
                for slot_pair in lab_slots:
                    if slot_pair[0] in lecture_slots and slot_pair[1] in lecture_slots:
                        selected_lab = random.choice(list(zip(self.lab_subject_entries, self.lab_faculty_entries)))
                        lab_subject, lab_faculties = selected_lab
                        selected_faculty = random.choice(lab_faculties)

                        if (selected_faculty not in faculty_schedule[day][slot_pair[0]]
                            and selected_faculty not in faculty_schedule[day][slot_pair[1]]):
                            timetables[division].at[day, slot_pair[0]] = f"{lab_subject} (Lab)\n{selected_faculty}"
                            timetables[division].at[day, slot_pair[1]] = f"{lab_subject} (Lab)\n{selected_faculty}"
                            faculty_schedule[day][slot_pair[0]].append(selected_faculty)
                            faculty_schedule[day][slot_pair[1]].append(selected_faculty)
                            lab_session_scheduled = True
                            break  # Schedule one lab session per day

                # Fill empty slots randomly
                for slot in all_slots:
                    if pd.isnull(timetables[division].at[day, slot]) and slot not in lunch_slots:
                        available_subjects = [subj for subj in self.subject_entries if subj not in used_subjects]
                        if available_subjects:
                            selected_subject = random.choice(available_subjects)
                            selected_faculty = random.choice(self.faculty_entries[self.subject_entries.index(selected_subject)])
                            if selected_faculty not in faculty_schedule[day][slot]:
                                timetables[division].at[day, slot] = f"{selected_subject}\n{selected_faculty}"
                                faculty_schedule[day][slot].append(selected_faculty)

                # Mark lunch slots with yellow color
                for lunch_slot in lunch_slots:
                    timetables[division].at[day, lunch_slot] = "Lunch Break"

        self.timetables = timetables

    def display_timetable(self):
        for division, timetable in self.timetables.items():
            print(f"\nDivision: {division}")
            print(timetable.to_string())

    def save_timetable_as_excel(self):
        if not self.timetables:
            print("Generate a timetable first!")
            return

        try:
            workbook = Workbook()
            for division, timetable in self.timetables.items():
                worksheet = workbook.create_sheet(title=division)

                # Set headers for days and times
                worksheet.append(['Day'] + list(timetable.columns))

                for r, day in enumerate(timetable.index, start=2):
                    row = [day]
                    for c in range(len(timetable.columns)):
                        cell_value = timetable.iat[r-2, c]
                        row.append(cell_value)
                    worksheet.append(row)

                for r in range(2, len(timetable.index) + 2):
                    for c in range(2, len(timetable.columns) + 2):
                        cell_value = worksheet.cell(row=r, column=c)
                        # Color coding
                        if "Lunch Break" in str(cell_value.value):
                            cell_value.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow for lunch
                        elif "(Lab)" in str(cell_value.value):
                            cell_value.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green for lab
                        elif cell_value.value is not None:
                            cell_value.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Light blue for subjects

                        # Apply borders
                        thin = Side(border_style="thin", color="000000")
                        cell_value.border = Border(left=thin, right=thin, top=thin, bottom=thin)

                # Apply borders to day and time headers
                for c in range(1, len(timetable.columns) + 1):
                    header_cell = worksheet.cell(row=1, column=c)
                    header_cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

                # Increase row height and column width
                for row in worksheet.iter_rows():
                    worksheet.row_dimensions[row[0].row].height = 30  # Increase row height
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter  # Get column letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.2
                    worksheet.column_dimensions[column_letter].width = adjusted_width  # Increase column width

            # Remove the default sheet created with the workbook
            if 'Sheet' in workbook.sheetnames:
                del workbook['Sheet']

            workbook.save("Timetables.xlsx")
            print("Timetable saved as 'Timetables.xlsx'.")
        except Exception as e:
            print(f"Failed to save timetable: {str(e)}")

if __name__ == "__main__":
    timetable_generator = TimetableGenerator()
    num_divisions, division_names, all_slots, lecture_slots, lunch_slots = timetable_generator.get_input()
    timetable_generator.generate_timetable(num_divisions, division_names, all_slots, lecture_slots, lunch_slots)
    timetable_generator.display_timetable()
    timetable_generator.save_timetable_as_excel()
